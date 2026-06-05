import os
import datetime as dt
import requests
from openpyxl import load_workbook

# === AYARLAR ===
CHAT_ID = int(os.getenv("TELEGRAM_CHAT_ID", "1677402217"))  # Sana gelecek özel mesaj
GROUP_CHAT_ID = -1003758241042  # ASKO YEMEK MENÜ grubuna gidecek mesaj

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
EXCEL_PATH = os.getenv("EXCEL_PATH", "Yemek_Listesi.xlsx")
SHEET_NAME = os.getenv("SHEET_NAME", "")
SEND_TOMORROW = os.getenv("SEND_TOMORROW", "false").lower() == "true"
# =================

MONTHS = {
    "Ocak": 1, "Şubat": 2, "Mart": 3, "Nisan": 4, "Mayıs": 5, "Haziran": 6,
    "Temmuz": 7, "Ağustos": 8, "Eylül": 9, "Ekim": 10, "Kasım": 11, "Aralık": 12
}

def parse_tr_date(s):
    if not s:
        return None
    parts = str(s).strip().split()
    if len(parts) != 3:
        return None
    try:
        return dt.date(int(parts[2]), MONTHS.get(parts[1]), int(parts[0]))
    except:
        return None

def get_menu(target_date):
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)

        if SHEET_NAME:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.active

    except Exception as e:
        return f"❗ Excel dosyası veya sayfası bulunamadı: {e}"

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Satır tamamen boşsa atla
        if not any(row):
            continue

        # --- KRİTİK SÜTUN DÜZELTMESİ ---
        # Gelen satırı listeye çevirip ilk 6 elemanını al (Ekstra çıkarıldı)
        veri = list(row)[:6]

        # Eğer satırda 6'dan az eleman varsa, eksikleri None ile tamamla
        veri += [None] * (6 - len(veri))

        t, gun, corba, ana, yard, tatli = veri

        d = parse_tr_date(t)
        if d == target_date:
            label = "Yarın" if SEND_TOMORROW else "Bugün"

            return (
                f"📌 {label} ({t} - {gun}) Menü:\n"
                f"🍲 Çorba: {corba}\n"
                f"🍽️ Yemek: {ana}\n"
                f"🥗 Yardımcı: {yard}\n"
                f"🍮 Tatlı/Meyve: {tatli}"
            )

    return f"❗ Menü bulunamadı: {target_date.strftime('%d.%m.%Y')}"

def send_telegram(msg, chat_id):
    if not TOKEN:
        print("HATA: TELEGRAM_BOT_TOKEN bulunamadı. GitHub Secrets'ı kontrol et.")
        return
        
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    try:
        r = requests.post(url, json={"chat_id": chat_id, "text": msg}, timeout=20)
        r.raise_for_status()
        print(f"Mesaj başarıyla gönderildi (ID: {chat_id})")
    except Exception as e:
        print(f"Mesaj gönderilirken hata oluştu (ID: {chat_id}): {e}")

if __name__ == "__main__":
    today = dt.date.today()
    target = today + dt.timedelta(days=1) if SEND_TOMORROW else today

    # Haftasonu: Cumartesi=5, Pazar=6 -> mesaj yok
    if target.weekday() >= 5:
        print("Haftasonu -> Mesaj gönderilmedi.")
        raise SystemExit(0)

    msg = get_menu(target)
    
    # 1. Sana (Admin) her durumda mesaj gitsin (hata olsa bile haberin olsun)
    send_telegram(msg, CHAT_ID)
    
    # 2. Gruba sadece menü BAŞARIYLA bulunduysa mesaj gitsin. 
    # (Resmi tatillerde Excel'de o gün yoksa gruba "Menü bulunamadı" diye boşuna mesaj atmaz)
    if "❗" not in msg:
        send_telegram(msg, GROUP_CHAT_ID)
    else:
        print("Hata veya tatil günü olduğu için gruba mesaj atılmadı.")
